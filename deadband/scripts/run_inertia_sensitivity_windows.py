#!/usr/bin/env python3
"""Representative-window inertia sensitivity for deadband candidates.

This script scales synchronous-generator inertia in a copied dynamic case and
evaluates the uniform baseline and the selected heterogeneous candidate on the
same Stage-1 hot-start windows. It is a lightweight robustness screen rather
than a full multi-system full-day validation.
"""

from __future__ import annotations

import argparse
import shutil
import warnings
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import openpyxl
import pandas as pd

import run_dispatch_tds as rdt
from compare_dispatch_pair_hotstart import AGC_ALLOCATION_HEADROOM, AGC_ANTI_WINDUP_FREEZE
from sweep_deadband_phase1_windows import (
    DEFAULT_WINDOWS,
    DeadbandCombo,
    WindowSpec,
    compute_trace_metrics,
    ordered_dispatch_labels,
    run_window_trace,
    summarize_combo,
)

COMPLEX_WARNING = getattr(getattr(np, "exceptions", object()), "ComplexWarning", None)
if COMPLEX_WARNING is not None:
    warnings.filterwarnings("ignore", category=COMPLEX_WARNING)
warnings.filterwarnings("ignore", message="Casting complex values to real discards")


def parse_window(text: str) -> WindowSpec:
    first, second = [part.strip() for part in text.split(",", 1)]
    return WindowSpec(first, second)


def scale_inertia_case(base_case: Path, out_case: Path, multiplier: float) -> dict[str, object]:
    if abs(float(multiplier) - 1.0) < 1e-12:
        return {
            "case_path": str(base_case.resolve()),
            "inertia_multiplier": float(multiplier),
            "scaled_rows": 0,
            "sum_M_before": float("nan"),
            "sum_M_after": float("nan"),
        }

    out_case.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_case, out_case)
    wb = openpyxl.load_workbook(out_case)
    scaled_rows = 0
    sum_before = 0.0
    sum_after = 0.0
    for sheet_name in ("GENROU", "GENROE", "GENSAL", "GENSAE", "GENCLS"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        if "M" not in header:
            continue
        col = header.index("M") + 1
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value in (None, ""):
                continue
            old = float(value)
            new = old * float(multiplier)
            ws.cell(row=row, column=col, value=new)
            scaled_rows += 1
            sum_before += old
            sum_after += new
    wb.save(out_case)
    return {
        "case_path": str(out_case.resolve()),
        "inertia_multiplier": float(multiplier),
        "scaled_rows": int(scaled_rows),
        "sum_M_before": float(sum_before),
        "sum_M_after": float(sum_after),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dispatch-dir", type=Path, required=True)
    parser.add_argument("--results-dir", type=Path, required=True)
    parser.add_argument("--curve-file", type=Path, default=rdt.DEFAULT_CURVE_FILE)
    parser.add_argument("--dyn-case", type=Path, required=True)
    parser.add_argument("--stable-dyn-case", type=Path, default=rdt.DEFAULT_STABLE_DYN_CASE)
    parser.add_argument("--inertia-multiplier", type=float, nargs="+", default=[0.5, 0.75, 1.0, 1.25])
    parser.add_argument("--window", action="append", type=parse_window, default=None)
    parser.add_argument("--dispatch-interval", type=int, default=900)
    parser.add_argument("--agc-interval", type=int, default=4)
    parser.add_argument("--kp", type=float, default=0.1)
    parser.add_argument("--ki", type=float, default=0.002)
    parser.add_argument("--wind-pref-alpha", type=float, default=0.98)
    parser.add_argument("--solar-pref-alpha", type=float, default=0.98)
    parser.add_argument("--disable-der-agc", action="store_true")
    parser.add_argument("--disable-pvd-agc", action="store_true")
    parser.add_argument("--disable-esd-agc", action="store_true")
    parser.add_argument("--agc-allocation-mode", default=AGC_ALLOCATION_HEADROOM)
    parser.add_argument("--agc-anti-windup-mode", default=AGC_ANTI_WINDUP_FREEZE)
    parser.add_argument("--agc-gov-output-ramp-frac-pmax-per-min", type=float, default=0.0)
    parser.add_argument("--agc-dg-output-ramp-frac-pmax-per-min", type=float, default=0.0)
    parser.add_argument("--init-mode", choices=("dispatch", "first"), default="first")
    parser.add_argument("--governor-target-schedule", default="ramp_limited_basepoint")
    parser.add_argument("--governor-basepoint-ramp-floor-frac-pmax-per-min", type=float, default=0.005)
    parser.add_argument("--governor-basepoint-ramp-gap-factor", type=float, default=1.25)
    parser.add_argument("--max-abs-hz-threshold", type=float, default=0.10)
    parser.add_argument("--share-abs-gt-0p05-threshold", type=float, default=0.05)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rdt.andes.config_logger(stream_level=30)
    args.results_dir.mkdir(parents=True, exist_ok=True)

    windows = args.window or [parse_window(text) for text in DEFAULT_WINDOWS]
    curve = rdt.load_curve(args.curve_file)
    available_labels = ordered_dispatch_labels(args.dispatch_dir)
    combos = {
        "uniform": DeadbandCombo(0.036, 0.036, 0.036),
        "best": DeadbandCombo(0.036, 0.025, 0.015),
    }

    runner_args = SimpleNamespace(
        dispatch_dir=args.dispatch_dir,
        dispatch_interval=args.dispatch_interval,
        agc_interval=args.agc_interval,
        kp=args.kp,
        ki=args.ki,
        wind_pref_alpha=args.wind_pref_alpha,
        solar_pref_alpha=args.solar_pref_alpha,
        disable_der_agc=args.disable_der_agc,
        disable_pvd_agc=args.disable_pvd_agc,
        disable_esd_agc=args.disable_esd_agc,
        agc_allocation_mode=args.agc_allocation_mode,
        agc_anti_windup_mode=args.agc_anti_windup_mode,
        agc_gov_output_ramp_frac_pmax_per_min=args.agc_gov_output_ramp_frac_pmax_per_min,
        agc_dg_output_ramp_frac_pmax_per_min=args.agc_dg_output_ramp_frac_pmax_per_min,
        init_mode=args.init_mode,
        governor_target_schedule=args.governor_target_schedule,
        governor_basepoint_ramp_floor_frac_pmax_per_min=args.governor_basepoint_ramp_floor_frac_pmax_per_min,
        governor_basepoint_ramp_gap_factor=args.governor_basepoint_ramp_gap_factor,
        max_abs_hz_threshold=args.max_abs_hz_threshold,
        share_abs_gt_0p05_threshold=args.share_abs_gt_0p05_threshold,
    )

    case_rows: list[dict[str, object]] = []
    window_rows: list[dict[str, object]] = []
    combo_rows: list[dict[str, object]] = []
    for multiplier in args.inertia_multiplier:
        tag = f"H{int(round(float(multiplier) * 100)):03d}"
        case_path = args.results_dir / "cases" / f"{args.stable_dyn_case.stem}_{tag}.xlsx"
        case_meta = scale_inertia_case(args.stable_dyn_case, case_path, float(multiplier))
        case_rows.append(case_meta)
        dyn_case = Path(case_meta["case_path"])
        for case_label, combo in combos.items():
            traces: list[pd.DataFrame] = []
            rows_for_combo: list[dict[str, object]] = []
            for window in windows:
                row: dict[str, object] = {
                    "inertia_multiplier": float(multiplier),
                    "case": case_label,
                    "combo_id": combo.combo_id,
                    "window": window.name,
                    "failed": 0,
                }
                try:
                    trace = run_window_trace(
                        curve=curve,
                        dyn_case=dyn_case,
                        window=window,
                        combo=combo,
                        args=runner_args,
                        available_labels=available_labels,
                    )
                    metrics = compute_trace_metrics(trace)
                    row.update(metrics)
                    trace_path = args.results_dir / f"{tag}_{case_label}_{window.name}_trace.csv"
                    trace.to_csv(trace_path, index=False)
                    row["trace_csv"] = str(trace_path)
                    traces.append(trace)
                except Exception as exc:
                    row["failed"] = 1
                    row["error"] = str(exc)
                    print(f"{tag} {case_label} {window.name} failed: {exc}")
                window_rows.append(row)
                rows_for_combo.append(row)
            summary = summarize_combo(combo, rows_for_combo, traces, runner_args)
            summary["inertia_multiplier"] = float(multiplier)
            summary["case"] = case_label
            combo_rows.append(summary)
            print(
                f"{tag} {case_label}: mean_abs={summary.get('mean_abs_hz', np.nan):.5f} "
                f"tail={summary.get('share_abs_gt_0p05', np.nan):.2%} "
                f"EM36={summary.get('edge_mass_36', np.nan):.2%} "
                f"failed={summary.get('failed_windows', 0)}"
            )

    pd.DataFrame(case_rows).to_csv(args.results_dir / "inertia_case_manifest.csv", index=False)
    pd.DataFrame(window_rows).to_csv(args.results_dir / "inertia_window_metrics.csv", index=False)
    combo_df = pd.DataFrame(combo_rows)
    combo_df.to_csv(args.results_dir / "inertia_combo_summary.csv", index=False)

    wide_rows = []
    for multiplier, sub in combo_df.groupby("inertia_multiplier"):
        if {"uniform", "best"}.issubset(set(sub["case"])):
            base = sub[sub["case"] == "uniform"].iloc[0]
            best = sub[sub["case"] == "best"].iloc[0]
            out = {"inertia_multiplier": float(multiplier)}
            for metric in ("mean_abs_hz", "share_abs_gt_0p036", "share_abs_gt_0p05", "max_abs_hz", "edge_mass_36"):
                out[f"uniform_{metric}"] = float(base[metric])
                out[f"best_{metric}"] = float(best[metric])
                out[f"delta_{metric}"] = float(best[metric] - base[metric])
                out[f"relative_{metric}"] = float((best[metric] - base[metric]) / base[metric]) if float(base[metric]) else np.nan
            wide_rows.append(out)
    pd.DataFrame(wide_rows).to_csv(args.results_dir / "inertia_sensitivity_wide.csv", index=False)
    print(f"summary_csv={args.results_dir / 'inertia_combo_summary.csv'}")


if __name__ == "__main__":
    main()
