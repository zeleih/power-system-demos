#!/usr/bin/env python3
"""
Run a deadband-demo dispatch interval through ANDES TDS.

The primary entrypoint is a dispatch JSON produced from the deadband demo
workflow, but the script can still recompute a dispatch from AMS when
``--dispatch-json`` is omitted.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
CASES = ROOT / "cases"
RESULTS = ROOT / "results"


def find_workspace(root: Path) -> Path | None:
    """
    Locate a workspace containing sibling ``andes`` and ``ams`` source trees.

    This keeps the demo runnable both inside the historical
    ``openandes/demo/demo/deadband`` layout and from a standalone export such as
    the ``deadband2`` repository.
    """
    candidates: list[Path] = []

    env_workspace = os.environ.get("OPENANDES_WORKSPACE")
    if env_workspace:
        candidates.append(Path(env_workspace).expanduser())

    resolved = root.resolve()
    for parent in (resolved, *resolved.parents):
        candidates.append(parent)
        candidates.append(parent / "openandes")

    seen: set[Path] = set()
    for candidate in candidates:
        try:
            candidate = candidate.resolve()
        except FileNotFoundError:
            continue

        if candidate in seen:
            continue
        seen.add(candidate)

        if (candidate / "andes").exists() and (candidate / "ams").exists():
            return candidate

    return None


WORKSPACE = find_workspace(ROOT)

# Prefer the local ANDES/AMS source trees so demo scripts pick up the
# compatibility patches in this workspace without requiring PYTHONPATH.
if WORKSPACE is not None:
    for src_root in (WORKSPACE / "andes", WORKSPACE / "ams"):
        src_str = str(src_root)
        if src_root.exists() and src_str not in sys.path:
            sys.path.insert(0, src_str)

import andes
from andes.thirdparty.npfunc import safe_div

DEFAULT_OPF_CASE = CASES / "IL200_opf2.xlsx"
DEFAULT_DYN_CASE = CASES / "IL200_dyn_db2.xlsx"
DEFAULT_CURVE_FILE = CASES / "CurveInterp.csv"
AGC_ALLOCATION_FIXED = "fixed_capacity"
AGC_ALLOCATION_HEADROOM = "headroom_aware"
AGC_ALLOCATION_MODES = (AGC_ALLOCATION_FIXED, AGC_ALLOCATION_HEADROOM)
AGC_ESD_SOC_MARGIN = 0.05
DEFAULT_STABLE_DYN_CASE = CASES / "IL200_dyn_db2_stable.xlsx"
DEFAULT_WIND_PREFIXES = ("WT_",)
DEFAULT_SOLAR_PREFIXES = ("PV_",)


@dataclass
class DispatchRecord:
    hour: int
    dispatch: int
    load: float
    wind: float
    solar: float
    gen: list
    pg: list
    qg: list
    pd: list
    bus: list
    vBus: list
    aBus: list
    converged: bool = True
    obj: float = float("nan")

    @property
    def label(self) -> str:
        return f"h{self.hour}d{self.dispatch}"

    @classmethod
    def from_json(cls, path: Path) -> "DispatchRecord":
        return cls(**json.loads(path.read_text()))


def dispatch_pg_map(dispatch_record: DispatchRecord) -> dict[int, float]:
    """
    Map static-generator idx to dispatch active power.

    The JSON `gen` ordering comes from AMS ACOPF output and is not guaranteed to
    match `sa.StaticGen.get_all_idxes()`. Always align through the explicit
    generator idx list instead of assuming positional equality.
    """
    return {
        int(gen): float(pg)
        for gen, pg in zip(dispatch_record.gen, dispatch_record.pg)
    }


def dispatch_online_mask(
    static_gen_idx: Iterable[int],
    dispatch_record: DispatchRecord,
    threshold: float = 1e-4,
) -> np.ndarray:
    """
    Return the online/offline mask in ANDES StaticGen order.
    """
    pg_map = dispatch_pg_map(dispatch_record)
    return np.array(
        [1.0 if float(pg_map.get(int(idx), 0.0)) > threshold else 0.0 for idx in static_gen_idx],
        dtype=float,
    )


def disable_der_frequency_deadband(sa) -> list[dict[str, object]]:
    """
    Disable frequency deadband / droop for PVD1 and ESD1 devices.

    This is useful when isolating the contribution of conventional governors.
    """
    touched: list[dict[str, object]] = []
    for model_name in ("PVD1", "ESD1"):
        if not hasattr(sa, model_name):
            continue
        mdl = getattr(sa, model_name)
        if mdl.n == 0:
            continue
        idx = mdl.idx.v
        zeros = np.zeros(mdl.n, dtype=float)
        for field in ("fdbd", "fdbdu", "ddn"):
            if hasattr(mdl, field):
                mdl.set(src=field, idx=idx, attr="v", value=zeros)
        touched.append({"model": model_name, "count": int(mdl.n)})
    return touched


def apply_traditional_governor_deadband(sa, deadband_hz: float) -> list[dict[str, object]]:
    """
    Apply a symmetric speed-input deadband to conventional governors.
    """
    db_pu = float(deadband_hz) / float(sa.config.freq)
    touched: list[dict[str, object]] = []

    for model_name in ("TGOV1NDB", "TGOV1DB", "HYGOVDB"):
        if not hasattr(sa, model_name):
            continue
        mdl = getattr(sa, model_name)
        if mdl.n == 0:
            continue

        idx = mdl.idx.v
        mdl.set(src="dbL", idx=idx, attr="v", value=np.full(mdl.n, -db_pu, dtype=float))
        mdl.set(src="dbU", idx=idx, attr="v", value=np.full(mdl.n, db_pu, dtype=float))
        r_values = mdl.get(src="R", attr="v", idx=idx)
        touched.append({
            "model": model_name,
            "count": int(mdl.n),
            "deadband_hz": float(deadband_hz),
            "deadband_pu": float(db_pu),
            "R_runtime_min": float(np.min(r_values)),
            "R_runtime_max": float(np.max(r_values)),
            "R_runtime_mean": float(np.mean(r_values)),
        })

    return touched


def apply_traditional_governor_deadband_csv(
    sa,
    csv_path: Path,
    *,
    model_name: str = "TGOV1NDB",
) -> list[dict[str, object]]:
    """
    Apply per-governor deadband settings from a CSV scheme.

    Expected columns:
    - gov_idx
    - dbL_pu / dbU_pu

    Optional columns used for reporting:
    - group
    - deadband_mhz
    - pmax_pu
    """
    csv_path = Path(csv_path).resolve()
    df = pd.read_csv(csv_path)
    required = {"gov_idx"}
    if not required.issubset(df.columns):
        missing = ", ".join(sorted(required - set(df.columns)))
        raise ValueError(f"Deadband CSV missing required column(s): {missing}")

    if "dbL_pu" not in df.columns or "dbU_pu" not in df.columns:
        if "deadband_mhz" not in df.columns:
            raise ValueError(
                "Deadband CSV must contain either dbL_pu/dbU_pu or deadband_mhz columns"
            )
        db_pu = pd.to_numeric(df["deadband_mhz"], errors="raise").astype(float) / float(sa.config.freq) / 1000.0
        df = df.copy()
        df["dbL_pu"] = -db_pu
        df["dbU_pu"] = db_pu

    if not hasattr(sa, model_name):
        raise RuntimeError(f"Model {model_name} not found in ANDES system")
    mdl = getattr(sa, model_name)
    if mdl.n == 0:
        raise RuntimeError(f"Model {model_name} has no devices in this case")

    idx_all = [str(idx) for idx in mdl.idx.v]
    df = df.copy()
    df["gov_idx"] = df["gov_idx"].astype(str)
    df = df[df["gov_idx"].isin(idx_all)]

    idx_set = set(idx_all)
    csv_set = set(df["gov_idx"])
    missing = sorted(idx_set - csv_set)
    extra = sorted(csv_set - idx_set)
    if missing:
        raise ValueError(
            f"Deadband CSV does not cover all runtime {model_name} devices. Missing: {missing}"
        )
    if extra:
        raise ValueError(
            f"Deadband CSV contains unknown {model_name} devices for this case: {extra}"
        )

    ordered = df.set_index("gov_idx").loc[idx_all].reset_index()
    db_l = pd.to_numeric(ordered["dbL_pu"], errors="raise").to_numpy(dtype=float)
    db_u = pd.to_numeric(ordered["dbU_pu"], errors="raise").to_numpy(dtype=float)
    mdl.set(src="dbL", idx=mdl.idx.v, attr="v", value=db_l)
    mdl.set(src="dbU", idx=mdl.idx.v, attr="v", value=db_u)

    touched: list[dict[str, object]] = []
    total_capacity = None
    if "pmax_pu" in ordered.columns:
        total_capacity = float(pd.to_numeric(ordered["pmax_pu"], errors="raise").sum())

    group_key = "group" if "group" in ordered.columns else None
    if group_key is not None:
        grouped = ordered.groupby(group_key, dropna=False, sort=False)
    else:
        grouped = [("all", ordered)]

    for group_name, part in grouped:
        meta: dict[str, object] = {
            "model": model_name,
            "group": str(group_name),
            "count": int(len(part)),
            "csv_path": str(csv_path),
        }
        if "deadband_mhz" in part.columns:
            unique_mhz = sorted(set(float(v) for v in pd.to_numeric(part["deadband_mhz"], errors="raise")))
            if len(unique_mhz) == 1:
                meta["deadband_hz"] = unique_mhz[0] / 1000.0
                meta["deadband_mhz"] = unique_mhz[0]
        if total_capacity is not None and "pmax_pu" in part.columns:
            capacity = float(pd.to_numeric(part["pmax_pu"], errors="raise").sum())
            meta["capacity_pu"] = capacity
            meta["capacity_share"] = capacity / total_capacity if total_capacity > 0.0 else float("nan")
        touched.append(meta)

    return touched


def normalize_prefixes(prefixes: Iterable[str] | None, defaults: tuple[str, ...]) -> tuple[str, ...]:
    if prefixes is None:
        return defaults

    items = tuple(prefix for prefix in prefixes if prefix)
    return items or defaults


def _validate_deadband_override(deadband_hz: float | None, *, name: str) -> float | None:
    if deadband_hz is None:
        return None
    deadband_hz = float(deadband_hz)
    if deadband_hz < 0.0:
        raise ValueError(f"{name} must be >= 0, got {deadband_hz}")
    return deadband_hz


def _apply_symmetric_deadband(model, idx: list[int] | list[str], deadband_hz: float) -> int:
    if not idx:
        return 0
    n = len(idx)
    lower = np.full(n, -float(deadband_hz), dtype=float)
    upper = np.full(n, float(deadband_hz), dtype=float)
    model.set(src="fdbd", idx=idx, attr="v", value=lower)
    model.set(src="fdbdu", idx=idx, attr="v", value=upper)
    return n


def adapt_dyn_case(src: Path, dst: Path) -> Path:
    """
    Create a stable-style copy of the legacy dynamic case.

    The legacy deadband case uses ``PVD2`` / ``ESD2``. For migration, rename
    them to ``PVD1`` / ``ESD1`` and add the optional ``fdbdu`` column
    explicitly. A default of ``0.017`` preserves the historical upper
    deadband used by the legacy bi-directional models when the column was
    omitted.
    """

    def rename_sheet(wb: openpyxl.Workbook, old: str, new: str) -> None:
        if new in wb.sheetnames and old in wb.sheetnames:
            del wb[new]
        if old in wb.sheetnames:
            wb[old].title = new

    def ensure_column(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        after: str,
        name: str,
        value: float,
    ) -> None:
        headers = [cell.value for cell in ws[1]]
        if name in headers or after not in headers:
            return

        insert_idx = headers.index(after) + 2
        ws.insert_cols(insert_idx)
        ws.cell(row=1, column=insert_idx, value=name)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=insert_idx, value=value)

    wb = openpyxl.load_workbook(src)
    rename_sheet(wb, "PVD2", "PVD1")
    rename_sheet(wb, "ESD2", "ESD1")

    for sheet in ("PVD1", "ESD1"):
        if sheet in wb.sheetnames:
            ensure_column(wb[sheet], after="fdbd", name="fdbdu", value=0.017)

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    return dst


def make_sp(opf_case: Path) -> Any:
    import ams

    return ams.load(str(opf_case), setup=True, no_output=True, default_config=True)


def load_curve(curve_file: Path) -> pd.DataFrame:
    return pd.read_csv(curve_file)


def compute_dispatch(
    hour: int,
    dispatch: int,
    curve: pd.DataFrame,
    opf_case: Path,
    dispatch_interval: int,
    wind_pref_alpha: float = 1.0,
    solar_pref_alpha: float = 1.0,
) -> DispatchRecord:
    """
    Recompute one ACOPF dispatch interval using the demo notebook logic.
    """
    r0 = hour * 3600 + dispatch * dispatch_interval
    r1 = r0 + dispatch_interval

    sp = make_sp(opf_case)

    pq_idx = sp.PQ.idx.v
    p0 = sp.PQ.p0.v.copy()
    q0 = sp.PQ.q0.v.copy()
    stg = sp.StaticGen.get_all_idxes()
    stg_w2t, stg_pv, _ = sp.StaticGen.find_idx(
        keys="gentype",
        values=["W2", "PV", "ES"],
        allow_all=True,
    )
    p0_w2t = sp.StaticGen.get(src="p0", attr="v", idx=stg_w2t)
    p0_pv = sp.StaticGen.get(src="p0", attr="v", idx=stg_pv)
    wind_pref_alpha = validate_pref_alpha(wind_pref_alpha, name="wind_pref_alpha")
    solar_pref_alpha = validate_pref_alpha(solar_pref_alpha, name="solar_pref_alpha")

    load = curve["Load"].iloc[r0:r1].values.mean()
    sp.PQ.set(src="p0", idx=pq_idx, attr="v", value=load * p0)
    sp.PQ.set(src="q0", idx=pq_idx, attr="v", value=load * q0)

    psum = sp.PQ.p0.v.sum()
    solar = curve["PV"].iloc[r0:r1].values.mean()
    wind = curve["Wind"].iloc[r0:r1].values.mean()

    wind_sum = wind_pref_alpha * wind * p0_w2t.sum()
    solar_sum = solar_pref_alpha * solar * p0_pv.sum()
    if wind_sum + solar_sum > psum:
        dgen = wind_sum + solar_sum - psum
        dwind = dgen / (wind_sum + solar_sum) * wind_sum
        dsolar = dgen / (wind_sum + solar_sum) * solar_sum
        wind = safe_div(wind_sum - 1.05 * dwind, wind_sum)
        solar = safe_div(solar_sum - 1.05 * dsolar, solar_sum)

    sp.StaticGen.set(
        src="p0",
        idx=stg_w2t,
        attr="v",
        value=der_pref_from_available(der_available_from_curve(wind, p0_w2t), wind_pref_alpha),
    )
    sp.StaticGen.set(
        src="p0",
        idx=stg_pv,
        attr="v",
        value=der_pref_from_available(der_available_from_curve(solar, p0_pv), solar_pref_alpha),
    )

    pmax = sp.StaticGen.get(src="pmax", attr="v", idx=stg).copy()
    sp.StaticGen.set(src="pmax", idx=stg, attr="v", value=0.995 * pmax)

    sp.ACOPF.config.update(verbose=0, out_all=0)
    sp.ACOPF.update()
    sp.ACOPF.run()

    sp.StaticGen.set(src="pmax", idx=stg, attr="v", value=pmax)

    return DispatchRecord(
        hour=hour,
        dispatch=dispatch,
        load=float(load),
        wind=float(wind),
        solar=float(solar),
        gen=sp.ACOPF.pg.get_all_idxes(),
        pg=sp.ACOPF.pg.v.tolist(),
        qg=sp.ACOPF.qg.v.tolist(),
        pd=sp.ACOPF.pd.v.tolist(),
        bus=sp.ACOPF.vBus.get_all_idxes(),
        vBus=sp.ACOPF.vBus.v.tolist(),
        aBus=sp.ACOPF.aBus.v.tolist(),
        converged=bool(sp.ACOPF.converged),
        obj=float(sp.ACOPF.obj.v),
    )


def build_andes_link(sa: andes.system.System) -> pd.DataFrame:
    """
    Build the minimal generator link table needed for AGC without AMS.
    """
    stg_idx = sa.StaticGen.get_all_idxes()
    dg_idx = sa.DG.find_idx(keys="gen", values=stg_idx, allow_none=True)
    rg_idx = sa.RenGen.find_idx(keys="gen", values=stg_idx, allow_none=True)
    pvd1_idx = set(map(str, sa.PVD1.idx.v)) if hasattr(sa, "PVD1") and sa.PVD1.n else set()
    esd1_idx = set(map(str, sa.ESD1.idx.v)) if hasattr(sa, "ESD1") and sa.ESD1.n else set()

    syg_idx = sa.SynGen.get_all_idxes()
    syg_gen = sa.SynGen.get(src="gen", attr="v", idx=syg_idx)
    gov_idx = sa.TurbineGov.find_idx(keys="syn", values=syg_idx, allow_none=True)
    gov_map = {int(gen): gov for gen, gov in zip(syg_gen, gov_idx)}

    gammap = np.ones(len(stg_idx), dtype=float)
    for i, dg in enumerate(dg_idx):
        if dg:
            gammap[i] = float(sa.DG.get(src="gammap", attr="v", idx=dg))
        elif rg_idx[i]:
            gammap[i] = float(sa.RenGen.get(src="gammap", attr="v", idx=rg_idx[i]))

    dg_model: list[str | None] = []
    for dg in dg_idx:
        if not dg:
            dg_model.append(None)
            continue
        dg_name = str(dg)
        if dg_name in pvd1_idx:
            dg_model.append("PVD1")
        elif dg_name in esd1_idx:
            dg_model.append("ESD1")
        else:
            dg_model.append("OTHER")

    link = pd.DataFrame(
        {
            "stg_idx": stg_idx,
            "gov_idx": [gov_map.get(int(idx)) for idx in stg_idx],
            "dg_idx": dg_idx,
            "dg_model": dg_model,
            "rg_idx": rg_idx,
            "gammap": gammap,
        }
    )
    link["has_gov"] = link["gov_idx"].notna().astype(int)
    link["has_dg"] = link["dg_idx"].notna().astype(int)
    link["has_rg"] = link["rg_idx"].notna().astype(int)
    link[["agov", "adg", "arg"]] = 0.0
    return link


def configure_der_agc_participation(
    sa: andes.system.System,
    link: pd.DataFrame,
    *,
    enable_der_agc: bool,
    enable_pvd_agc: bool | None = None,
    enable_esd_agc: bool | None = None,
    enable_other_dg_agc: bool | None = None,
) -> pd.DataFrame:
    """
    Enable or disable DER participation in AGC without changing DER pref0.

    This keeps the DG/PVD1/ESD1 base power reference path intact and only
    controls whether AGC writes to DG.Pext0.
    """
    if enable_pvd_agc is None:
        enable_pvd_agc = enable_der_agc
    if enable_esd_agc is None:
        enable_esd_agc = enable_der_agc
    if enable_other_dg_agc is None:
        enable_other_dg_agc = enable_der_agc

    out = link.copy()
    dg_model = out["dg_model"].astype(object)
    dg_enabled = np.ones(len(out), dtype=bool)
    dg_enabled[(dg_model == "PVD1").to_numpy()] = bool(enable_pvd_agc)
    dg_enabled[(dg_model == "ESD1").to_numpy()] = bool(enable_esd_agc)
    dg_enabled[(dg_model == "OTHER").to_numpy()] = bool(enable_other_dg_agc)

    out["has_dg"] = np.asarray(out["has_dg"], dtype=int) * dg_enabled.astype(int)
    disabled_mask = np.asarray(link["has_dg"], dtype=int).astype(bool) & (~dg_enabled)
    out.loc[disabled_mask, "adg"] = 0.0

    if np.any(disabled_mask) and hasattr(sa, "DG") and sa.DG.n:
        disabled_dg_idx = [idx for idx in out.loc[disabled_mask, "dg_idx"].tolist() if pd.notna(idx)]
        if disabled_dg_idx:
            sa.DG.set(
                src="Pext0",
                idx=disabled_dg_idx,
                attr="v",
                value=np.zeros(len(disabled_dg_idx), dtype=float),
            )
    return out


def compute_agc_allocation_shares(
    sa: andes.system.System,
    link: pd.DataFrame,
    bf: np.ndarray,
    *,
    ace_raw: float,
    pext_max: np.ndarray,
    allocation_mode: str = AGC_ALLOCATION_HEADROOM,
) -> np.ndarray:
    """
    Return normalized AGC allocation shares aligned with ``link`` rows.

    ``fixed_capacity`` reproduces the historical static-capacity allocation.
    ``headroom_aware`` multiplies those same base weights by current regulation
    headroom in the requested direction before renormalizing.
    """
    shares = np.zeros(len(link), dtype=float)
    base = np.zeros(len(link), dtype=float)
    sn = np.asarray(sa.StaticGen.get(src="Sn", attr="v", idx=link["stg_idx"].tolist()), dtype=float)
    gammap = np.asarray(link["gammap"], dtype=float)
    dispatch_online = (np.asarray(bf, dtype=float) > 0.0).astype(float)
    gov_mask = np.asarray(link["has_gov"], dtype=bool)
    dg_mask = np.asarray(link["has_dg"], dtype=bool)
    dg_model = link["dg_model"].astype(object).to_numpy()

    # Conventional units still use dispatch pg>0 as the online test so AGC is
    # not assigned to thermal machines that are not running.
    base[gov_mask] = sn[gov_mask] * dispatch_online[gov_mask]

    # Non-storage DGs preserve the historical dispatch-online behavior.
    non_storage_dg_mask = dg_mask & (dg_model != "ESD1")
    base[non_storage_dg_mask] = sn[non_storage_dg_mask] * dispatch_online[non_storage_dg_mask]

    # Storage can be AGC-available while scheduled at pg=0, so gate it by
    # device status `u` instead of dispatch power.
    storage_mask = dg_mask & (dg_model == "ESD1")
    if np.any(storage_mask):
        esd_idx = link.loc[storage_mask, "dg_idx"].tolist()
        esd_u = np.asarray(sa.ESD1.get(src="u", attr="v", idx=esd_idx), dtype=float)
        base[storage_mask] = sn[storage_mask] * (esd_u > 0.5).astype(float)

    eligible = (np.asarray(link["has_gov"], dtype=int) + np.asarray(link["has_dg"], dtype=int)) > 0
    base = base * gammap * eligible.astype(float)

    total_base = float(base.sum())
    if total_base <= 0.0:
        return shares

    if allocation_mode != AGC_ALLOCATION_HEADROOM or abs(float(ace_raw)) <= 1e-12:
        return base / total_base

    direction_up = float(ace_raw) >= 0.0
    dynamic = np.zeros(len(link), dtype=float)

    if np.any(gov_mask):
        gov_idx = link.loc[gov_mask, "gov_idx"].tolist()
        gov_syn = sa.TurbineGov.get(src="syn", attr="v", idx=gov_idx)
        gov_gen = sa.SynGen.get(src="gen", attr="v", idx=gov_syn)
        gov_pmax = np.asarray(sa.StaticGen.get(src="pmax", attr="v", idx=gov_gen), dtype=float)
        gov_pmin = np.asarray(sa.StaticGen.get(src="pmin", attr="v", idx=gov_gen), dtype=float)
        gov_pref0 = np.asarray(sa.TurbineGov.get(src="pref0", attr="v", idx=gov_idx), dtype=float)
        gov_avail = gov_pmax - gov_pref0 if direction_up else gov_pref0 - gov_pmin
        dynamic[gov_mask] = np.maximum(0.0, gov_avail)

    if np.any(dg_mask):
        dg_idx = link.loc[dg_mask, "dg_idx"].tolist()
        dg_prev = np.asarray(sa.DG.get(src="Pext0", attr="v", idx=dg_idx), dtype=float)
        dg_upper = np.asarray(pext_max[sa.DG.idx2uid(dg_idx)], dtype=float)
        if direction_up:
            dg_avail = np.maximum(0.0, dg_upper - dg_prev)
        else:
            # Downward DG Pext0 is currently unbounded in this demo. Keep DGs
            # eligible without introducing an artificial hard lower limit.
            dg_avail = np.ones_like(dg_prev, dtype=float)

        storage_local_mask = (link.loc[dg_mask, "dg_model"].astype(object).to_numpy() == "ESD1")
        if np.any(storage_local_mask):
            esd_idx = [idx for idx, is_storage in zip(dg_idx, storage_local_mask) if is_storage]
            esd_uid = sa.ESD1.idx2uid(esd_idx)
            esd_u = np.asarray(sa.ESD1.u.v[esd_uid], dtype=float)
            if hasattr(sa.ESD1, "pIG_y") and getattr(sa.ESD1.pIG_y, "v", None) is not None and len(sa.ESD1.pIG_y.v):
                esd_soc = np.asarray(sa.ESD1.pIG_y.v[esd_uid], dtype=float)
            else:
                esd_soc = np.asarray(sa.ESD1.SOCinit.v[esd_uid], dtype=float)
            esd_socmin = np.asarray(sa.ESD1.SOCmin.v[esd_uid], dtype=float)
            esd_socmax = np.asarray(sa.ESD1.SOCmax.v[esd_uid], dtype=float)
            if direction_up:
                esd_available = ((esd_u > 0.5) & (esd_soc > (esd_socmin + AGC_ESD_SOC_MARGIN))).astype(float)
            else:
                esd_available = ((esd_u > 0.5) & (esd_soc < (esd_socmax - AGC_ESD_SOC_MARGIN))).astype(float)
            dg_avail[storage_local_mask] = esd_available

        dynamic[dg_mask] = dg_avail

    weighted = base * dynamic
    total_weighted = float(weighted.sum())
    if total_weighted <= 0.0:
        return base / total_base
    return weighted / total_weighted


def pvd1_gen_subsets(
    sa: andes.system.System,
    wind_prefixes: Iterable[str],
    solar_prefixes: Iterable[str],
) -> tuple[list[int], list[int]]:
    """
    Split PVD1 devices into wind and PV subsets using idx/name prefixes.
    """
    wind_prefixes = tuple(wind_prefixes)
    solar_prefixes = tuple(solar_prefixes)

    names = getattr(sa.PVD1, "name", None)
    name_values = names.v if names is not None else [None] * sa.PVD1.n

    wind: list[int] = []
    solar: list[int] = []

    for idx, name, gen in zip(sa.PVD1.idx.v, name_values, sa.PVD1.gen.v):
        labels = [str(idx)]
        if name is not None:
            labels.append(str(name))

        if any(label.startswith(prefix) for label in labels for prefix in wind_prefixes):
            wind.append(int(gen))
        if any(label.startswith(prefix) for label in labels for prefix in solar_prefixes):
            solar.append(int(gen))

    if not wind or not solar:
        sample = ", ".join(map(str, sa.PVD1.idx.v[:10]))
        raise ValueError(
            "Unable to classify PVD1 devices from prefixes. "
            f"wind_prefixes={wind_prefixes}, solar_prefixes={solar_prefixes}, "
            f"sample_idx=[{sample}]"
        )

    return wind, solar


def apply_resource_deadband_overrides(
    sa: andes.system.System,
    *,
    wind_prefixes: Iterable[str],
    solar_prefixes: Iterable[str],
    wind_deadband_hz: float | None = None,
    solar_deadband_hz: float | None = None,
    esd_deadband_hz: float | None = None,
) -> dict[str, object]:
    """
    Apply per-resource deadband overrides while preserving the existing ddn.

    These overrides are intended for phase-1 deadband studies where wind, PV,
    and storage deadbands are scanned independently on top of a fixed dynamic
    case and dispatch baseline.
    """
    wind_deadband_hz = _validate_deadband_override(wind_deadband_hz, name="wind_deadband_hz")
    solar_deadband_hz = _validate_deadband_override(solar_deadband_hz, name="solar_deadband_hz")
    esd_deadband_hz = _validate_deadband_override(esd_deadband_hz, name="esd_deadband_hz")

    meta: dict[str, object] = {
        "wind_deadband_hz": wind_deadband_hz,
        "solar_deadband_hz": solar_deadband_hz,
        "esd_deadband_hz": esd_deadband_hz,
        "configured_wind_pvd1_count": 0,
        "configured_solar_pvd1_count": 0,
        "configured_esd1_count": 0,
    }

    if hasattr(sa, "PVD1") and sa.PVD1.n and (wind_deadband_hz is not None or solar_deadband_hz is not None):
        stg_w2t, stg_pv = pvd1_gen_subsets(sa, wind_prefixes, solar_prefixes)
        if wind_deadband_hz is not None:
            pvd1_w2t = list(sa.PVD1.find_idx(keys="gen", values=stg_w2t))
            meta["configured_wind_pvd1_count"] = _apply_symmetric_deadband(sa.PVD1, pvd1_w2t, wind_deadband_hz)
        if solar_deadband_hz is not None:
            pvd1_pv = list(sa.PVD1.find_idx(keys="gen", values=stg_pv))
            meta["configured_solar_pvd1_count"] = _apply_symmetric_deadband(sa.PVD1, pvd1_pv, solar_deadband_hz)

    if hasattr(sa, "ESD1") and sa.ESD1.n and esd_deadband_hz is not None:
        esd_idx = list(sa.ESD1.idx.v)
        meta["configured_esd1_count"] = _apply_symmetric_deadband(sa.ESD1, esd_idx, esd_deadband_hz)

    return meta


def validate_curve_window(curve: pd.DataFrame, dispatch_record: DispatchRecord, duration_seconds: int) -> None:
    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds
    r1 = r0 + duration_seconds
    if r1 > len(curve):
        raise ValueError(
            f"Curve data only has {len(curve)} samples but {dispatch_record.label} "
            f"needs samples [{r0}, {r1})."
        )


def resolve_initial_profile(
    curve: pd.DataFrame,
    dispatch_record: DispatchRecord,
    duration_seconds: int,
    init_mode: str,
) -> tuple[float, float, float]:
    """
    Resolve the TDS starting load / wind / solar point.

    ``dispatch`` preserves the historical behavior of initializing from the
    dispatch-interval average. ``first`` uses the first curve sample in the
    interval.
    """
    if init_mode == "dispatch":
        return (
            float(dispatch_record.load),
            float(dispatch_record.wind),
            float(dispatch_record.solar),
        )

    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds
    return (
        float(curve["Load"].iloc[r0]),
        float(curve["Wind"].iloc[r0]),
        float(curve["PV"].iloc[r0]),
    )


def validate_pref_alpha(alpha: float, *, name: str) -> float:
    alpha = float(alpha)
    if not (0.0 <= alpha <= 1.0):
        raise ValueError(f"{name} must be within [0, 1], got {alpha}")
    return alpha


def der_available_from_curve(scale: float, p0_template: np.ndarray) -> np.ndarray:
    return float(scale) * np.asarray(p0_template, dtype=float)


def der_pref_from_available(available: np.ndarray, alpha: float) -> np.ndarray:
    return float(alpha) * np.asarray(available, dtype=float)


def run_tds(
    dispatch_record: DispatchRecord,
    curve: pd.DataFrame,
    dyn_case: Path,
    duration_seconds: int,
    agc_interval: int,
    kp: float,
    ki: float,
    wind_prefixes: Iterable[str],
    solar_prefixes: Iterable[str],
    agc_allocation_mode: str = AGC_ALLOCATION_HEADROOM,
    init_mode: str = "first",
    enable_der_agc: bool = True,
    enable_pvd_agc: bool | None = None,
    enable_esd_agc: bool | None = None,
    wind_pref_alpha: float = 1.0,
    solar_pref_alpha: float = 1.0,
    wind_deadband_hz: float | None = None,
    solar_deadband_hz: float | None = None,
    esd_deadband_hz: float | None = None,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Run one dispatch interval and return time and ACE frequency deviation.
    """
    validate_curve_window(curve, dispatch_record, duration_seconds)

    sa = andes.load(str(dyn_case), setup=False, no_output=True, default_config=True)

    # Record a direct frequency trace for plotting.
    sa.add("Output", dict(model="ACEc", varname="f"))

    sa.setup()

    link = configure_der_agc_participation(
        sa,
        build_andes_link(sa),
        enable_der_agc=enable_der_agc,
        enable_pvd_agc=enable_pvd_agc,
        enable_esd_agc=enable_esd_agc,
    )

    pq_idx = sa.PQ.idx.v
    stg = sa.StaticGen.get_all_idxes()
    stg_w2t, stg_pv = pvd1_gen_subsets(sa, wind_prefixes, solar_prefixes)
    p0_w2t = sa.StaticGen.get(src="p0", attr="v", idx=stg_w2t)
    p0_pv = sa.StaticGen.get(src="p0", attr="v", idx=stg_pv)
    pvd1_w2t = sa.PVD1.find_idx(keys="gen", values=stg_w2t)
    pvd1_pv = sa.PVD1.find_idx(keys="gen", values=stg_pv)
    wind_pref_alpha = validate_pref_alpha(wind_pref_alpha, name="wind_pref_alpha")
    solar_pref_alpha = validate_pref_alpha(solar_pref_alpha, name="solar_pref_alpha")
    apply_resource_deadband_overrides(
        sa,
        wind_prefixes=wind_prefixes,
        solar_prefixes=solar_prefixes,
        wind_deadband_hz=wind_deadband_hz,
        solar_deadband_hz=solar_deadband_hz,
        esd_deadband_hz=esd_deadband_hz,
    )

    sap0 = sa.PQ.p0.v.copy()
    saq0 = sa.PQ.q0.v.copy()

    sa.StaticGen.set(src="p0", idx=dispatch_record.gen, attr="v", value=dispatch_record.pg)
    sa.Bus.set(src="v0", idx=dispatch_record.bus, attr="v", value=dispatch_record.vBus)
    sa.Bus.set(src="a0", idx=dispatch_record.bus, attr="v", value=dispatch_record.aBus)

    pv_bus = sa.PV.bus.v
    slack_bus = sa.Slack.bus.v
    v_pv = sa.Bus.get(src="v0", attr="v", idx=pv_bus)
    a_slack = sa.Bus.get(src="a0", attr="v", idx=slack_bus)
    sa.PV.set(src="v0", idx=sa.PV.idx.v, attr="v", value=v_pv)
    sa.Slack.set(src="a0", idx=sa.Slack.idx.v, attr="v", value=a_slack)

    stg_on = dispatch_online_mask(stg, dispatch_record)
    sn = sa.StaticGen.get(src="Sn", attr="v", idx=stg)
    bf = stg_on * sn / (stg_on * sn).sum()

    sa.PQ.config.p2p = 1
    sa.PQ.config.q2q = 1
    sa.PQ.config.p2z = 0
    sa.PQ.config.q2z = 0
    sa.PQ.pq2z = 0

    sa.TDS.config.criteria = 0
    sa.TDS.config.no_tqdm = True

    init_load, init_wind, init_solar = resolve_initial_profile(
        curve=curve,
        dispatch_record=dispatch_record,
        duration_seconds=duration_seconds,
        init_mode=init_mode,
    )

    init_wind_pavail = der_available_from_curve(init_wind, p0_w2t)
    init_solar_pavail = der_available_from_curve(init_solar, p0_pv)
    init_wind_pref = der_pref_from_available(init_wind_pavail, wind_pref_alpha)
    init_solar_pref = der_pref_from_available(init_solar_pavail, solar_pref_alpha)

    sa.PQ.set(src="p0", idx=pq_idx, attr="v", value=init_load * sap0)
    sa.PQ.set(src="q0", idx=pq_idx, attr="v", value=init_load * saq0)
    sa.StaticGen.set(src="p0", idx=stg_w2t, attr="v", value=init_wind_pref)
    sa.StaticGen.set(src="p0", idx=stg_pv, attr="v", value=init_solar_pref)
    if pvd1_w2t:
        sa.PVD1.set(src="pref0", idx=pvd1_w2t, attr="v", value=init_wind_pref)
        sa.PVD1.set(src="pavail0", idx=pvd1_w2t, attr="v", value=init_wind_pavail)
    if pvd1_pv:
        sa.PVD1.set(src="pref0", idx=pvd1_pv, attr="v", value=init_solar_pref)
        sa.PVD1.set(src="pavail0", idx=pvd1_pv, attr="v", value=init_solar_pavail)

    sa.PFlow.run()
    if sa.exit_code != 0:
        raise RuntimeError(f"PFlow failed with exit_code={sa.exit_code}")

    _ = sa.TDS.init()
    if sa.exit_code != 0:
        raise RuntimeError(f"TDS init failed with exit_code={sa.exit_code}")

    pext_max = 999 * np.ones(sa.DG.n)
    if hasattr(sa, "ESD1") and sa.ESD1.n:
        ess_uid = sa.DG.idx2uid(sa.ESD1.idx.v)
        pext_max[ess_uid] = 999

    ace_integral = 0.0
    ace_raw = 0.0
    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds
    t_snapshots = [0.0]
    f_snapshots = [float((sa.ACEc.f.v[0] - 1.0) * sa.config.freq)]

    for t in range(duration_seconds):
        shares = compute_agc_allocation_shares(
            sa,
            link,
            bf,
            ace_raw=ace_raw,
            pext_max=pext_max,
            allocation_mode=agc_allocation_mode,
        )
        link["agov"] = ace_raw * shares * np.asarray(link["has_gov"], dtype=float)
        link["adg"] = ace_raw * shares * np.asarray(link["has_dg"], dtype=float)
        link["arg"] = 0.0

        if t % agc_interval == 0 and t > 0:
            agov_to_set = {
                gov: agov for gov, agov in zip(link["gov_idx"], link["agov"]) if pd.notna(gov)
            }
            if agov_to_set:
                gov_idx = list(agov_to_set.keys())
                paux0_raw = np.array(list(agov_to_set.values()))
                gov_syn = sa.TurbineGov.get(src="syn", attr="v", idx=gov_idx)
                gov_gen = sa.SynGen.get(src="gen", attr="v", idx=gov_syn)
                gov_pmax = sa.StaticGen.get(src="pmax", attr="v", idx=gov_gen)
                gov_pmin = sa.StaticGen.get(src="pmin", attr="v", idx=gov_gen)
                gov_pref0 = sa.TurbineGov.get(src="pref0", attr="v", idx=gov_idx)
                gov_up = np.maximum(0.0, gov_pmax - gov_pref0)
                gov_dn = np.minimum(0.0, gov_pmin - gov_pref0)
                # Preserve the AGC command sign while respecting generator headroom.
                paux0 = np.where(
                    paux0_raw >= 0.0,
                    np.minimum(paux0_raw, gov_up),
                    np.maximum(paux0_raw, gov_dn),
                )
                sa.TurbineGov.set(src="paux0", idx=gov_idx, attr="v", value=paux0)

            adg_to_set = {dg: adg for dg, adg in zip(link["dg_idx"], link["adg"]) if pd.notna(dg)}
            if adg_to_set:
                dg_idx = list(adg_to_set.keys())
                pext0_raw = np.array(list(adg_to_set.values()))
                dg_uids = sa.DG.idx2uid(dg_idx)
                pext0 = np.minimum(pext0_raw, pext_max[dg_uids])
                sa.DG.set(src="Pext0", idx=dg_idx, attr="v", value=pext0)

        if t > 0:
            kload = curve["Load"].iloc[r0 + t]
            sa.PQ.set(src="Ppf", idx=sa.PQ.idx.v, attr="v", value=kload * sap0)
            sa.PQ.set(src="Qpf", idx=sa.PQ.idx.v, attr="v", value=kload * saq0)

            wind = curve["Wind"].iloc[r0 + t]
            wind_pavail = der_available_from_curve(wind, p0_w2t)
            wind_pref = der_pref_from_available(wind_pavail, wind_pref_alpha)
            sa.PVD1.set(src="pref0", idx=pvd1_w2t, attr="v", value=wind_pref)
            sa.PVD1.set(src="pavail0", idx=pvd1_w2t, attr="v", value=wind_pavail)

            solar = curve["PV"].iloc[r0 + t]
            solar_pavail = der_available_from_curve(solar, p0_pv)
            solar_pref = der_pref_from_available(solar_pavail, solar_pref_alpha)
            sa.PVD1.set(src="pref0", idx=pvd1_pv, attr="v", value=solar_pref)
            sa.PVD1.set(src="pavail0", idx=pvd1_pv, attr="v", value=solar_pavail)

            sa.TDS.config.tf = t
            sa.TDS.run()
            t_snapshots.append(float(sa.dae.t))
            f_snapshots.append(float((sa.ACEc.f.v[0] - 1.0) * sa.config.freq))

            ace_sum = sa.ACEc.ace.v.sum()
            ace_raw = -(kp * ace_sum + ki * ace_integral)
            ace_integral = ace_integral + ace_sum

        if sa.exit_code != 0:
            raise RuntimeError(f"TDS failed at t={t}s with exit_code={sa.exit_code}")

    t = np.asarray(sa.dae.ts.t).reshape(-1)
    f_pu = np.asarray(sa.dae.ts.get_data(sa.ACEc.f, a=[0])).reshape(-1)
    n = min(len(t), len(f_pu))
    if n > 0:
        f_dev_hz = (f_pu[:n] - 1.0) * sa.config.freq
        return t[:n], f_dev_hz

    return np.asarray(t_snapshots), np.asarray(f_snapshots)


def save_outputs(
    t: np.ndarray,
    f_dev_hz: np.ndarray,
    dispatch_record: DispatchRecord,
    out_dir: Path,
    label: str | None = None,
    save_plot: bool = True,
) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{label or dispatch_record.label}_frequency"

    csv_path = out_dir / f"{stem}.csv"
    png_path = out_dir / f"{stem}.png"

    pd.DataFrame({"time_s": t, "freq_dev_hz": f_dev_hz}).to_csv(csv_path, index=False)

    if save_plot:
        try:
            os.environ.setdefault("MPLBACKEND", "Agg")
            if "MPLCONFIGDIR" not in os.environ:
                _mplconfig = Path(os.environ.get("TMPDIR", "/tmp")) / "openandes-mpl"
                _mplconfig.mkdir(parents=True, exist_ok=True)
                os.environ["MPLCONFIGDIR"] = str(_mplconfig)

            import matplotlib.pyplot as plt

            fig, ax = plt.subplots(figsize=(9, 4.8))
            ax.plot(t, f_dev_hz, color="#0f5c78", linewidth=1.4)
            ax.axhline(0.0, color="#777777", linewidth=0.8, linestyle="--")
            ax.set_title(f"Deadband Demo Frequency Deviation ({label or dispatch_record.label})")
            ax.set_xlabel("Time [s]")
            ax.set_ylabel("Frequency deviation [Hz]")
            ax.grid(True, alpha=0.25)
            fig.tight_layout()
            fig.savefig(png_path, dpi=180)
            plt.close(fig)
        except Exception as exc:
            print(f"warning: failed to save plot {png_path}: {exc}", file=sys.stderr)

    return csv_path, png_path


def write_dispatch_json(
    dispatch_record: DispatchRecord,
    out_dir: Path,
    label: str | None = None,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{label or dispatch_record.label}_dispatch.json"
    path.write_text(json.dumps(dispatch_record.__dict__, indent=2))
    return path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dispatch-json", type=Path, default=None,
                        help="Existing dispatch JSON to replay through TDS.")
    parser.add_argument("--hour", type=int, default=13,
                        help="Dispatch hour used when recomputing from AMS.")
    parser.add_argument("--dispatch", type=int, default=2,
                        help="Dispatch interval used when recomputing from AMS.")
    parser.add_argument("--label", type=str, default=None,
                        help="Output label. Defaults to h<hour>d<dispatch>.")
    parser.add_argument("--opf-case", type=Path, default=DEFAULT_OPF_CASE)
    parser.add_argument("--dyn-case", type=Path, default=DEFAULT_DYN_CASE)
    parser.add_argument("--stable-dyn-case", type=Path, default=DEFAULT_STABLE_DYN_CASE)
    parser.add_argument("--curve-file", type=Path, default=DEFAULT_CURVE_FILE)
    parser.add_argument("--results-dir", type=Path, default=RESULTS)
    parser.add_argument("--duration-seconds", type=int, default=900)
    parser.add_argument("--agc-interval", type=int, default=4)
    parser.add_argument("--kp", type=float, default=0.03)
    parser.add_argument("--ki", type=float, default=0.01)
    parser.add_argument(
        "--wind-pref-alpha",
        type=float,
        default=1.0,
        help="Scale wind pref0 relative to wind pavail0. 1 keeps pref0=pavail0.",
    )
    parser.add_argument(
        "--solar-pref-alpha",
        type=float,
        default=1.0,
        help="Scale solar pref0 relative to solar pavail0. 1 keeps pref0=pavail0.",
    )
    parser.add_argument(
        "--agc-allocation-mode",
        choices=AGC_ALLOCATION_MODES,
        default=AGC_ALLOCATION_HEADROOM,
    )
    parser.add_argument("--disable-der-agc", action="store_true")
    parser.add_argument(
        "--disable-pvd-agc",
        action="store_true",
        help="Do not allocate AGC Pext0 commands to wind/PV PVD1 devices.",
    )
    parser.add_argument(
        "--disable-esd-agc",
        action="store_true",
        help="Do not allocate AGC Pext0 commands to ESD1 storage devices.",
    )
    parser.add_argument("--init-mode", choices=("dispatch", "first"),
                        default="first",
                        help="TDS initialization profile: dispatch average or a curve sample.")
    parser.add_argument("--wind-prefix", action="append", default=None,
                        help="PVD1 idx/name prefix for wind units. Repeatable.")
    parser.add_argument("--solar-prefix", action="append", default=None,
                        help="PVD1 idx/name prefix for solar units. Repeatable.")
    parser.add_argument(
        "--wind-deadband-hz",
        type=float,
        default=None,
        help="Override wind PVD1 frequency deadband in Hz. Leaves ddn unchanged.",
    )
    parser.add_argument(
        "--solar-deadband-hz",
        type=float,
        default=None,
        help="Override solar PVD1 frequency deadband in Hz. Leaves ddn unchanged.",
    )
    parser.add_argument(
        "--esd-deadband-hz",
        type=float,
        default=None,
        help="Override ESD1 frequency deadband in Hz. Leaves ddn unchanged.",
    )
    parser.add_argument("--save-plot", dest="save_plot", action="store_true")
    parser.add_argument("--no-save-plot", dest="save_plot", action="store_false")
    parser.set_defaults(save_plot=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    andes.config_logger(stream_level=30)

    curve = load_curve(args.curve_file)
    dyn_case = adapt_dyn_case(args.dyn_case, args.stable_dyn_case)
    wind_prefixes = normalize_prefixes(args.wind_prefix, DEFAULT_WIND_PREFIXES)
    solar_prefixes = normalize_prefixes(args.solar_prefix, DEFAULT_SOLAR_PREFIXES)

    if args.dispatch_json is not None:
        dispatch_record = DispatchRecord.from_json(args.dispatch_json)
    else:
        import ams

        ams.config_logger(stream_level=50)
        dispatch_record = compute_dispatch(
            args.hour,
            args.dispatch,
            curve,
            args.opf_case,
            args.duration_seconds,
            wind_pref_alpha=args.wind_pref_alpha,
            solar_pref_alpha=args.solar_pref_alpha,
        )

    if not dispatch_record.converged:
        raise RuntimeError(f"Dispatch {dispatch_record.label} did not converge")

    label = args.label or dispatch_record.label
    t, f_dev_hz = run_tds(
        dispatch_record=dispatch_record,
        curve=curve,
        dyn_case=dyn_case,
        duration_seconds=args.duration_seconds,
        agc_interval=args.agc_interval,
        kp=args.kp,
        ki=args.ki,
        agc_allocation_mode=args.agc_allocation_mode,
        wind_prefixes=wind_prefixes,
        solar_prefixes=solar_prefixes,
        init_mode=args.init_mode,
        enable_der_agc=not args.disable_der_agc,
        enable_pvd_agc=not args.disable_pvd_agc,
        enable_esd_agc=not args.disable_esd_agc,
        wind_pref_alpha=args.wind_pref_alpha,
        solar_pref_alpha=args.solar_pref_alpha,
        wind_deadband_hz=args.wind_deadband_hz,
        solar_deadband_hz=args.solar_deadband_hz,
        esd_deadband_hz=args.esd_deadband_hz,
    )
    dispatch_json = write_dispatch_json(dispatch_record, args.results_dir, label=label)
    csv_path, png_path = save_outputs(
        t,
        f_dev_hz,
        dispatch_record,
        args.results_dir,
        label=label,
        save_plot=args.save_plot,
    )

    print(f"dispatch_json={dispatch_json}")
    print(f"freq_csv={csv_path}")
    print(f"freq_plot={png_path}")
    print(f"freq_dev_min_hz={float(f_dev_hz.min())}")
    print(f"freq_dev_max_hz={float(f_dev_hz.max())}")
    print(f"samples={len(t)}")


if __name__ == "__main__":
    main()
