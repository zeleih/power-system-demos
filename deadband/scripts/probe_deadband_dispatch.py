#!/usr/bin/env python3
"""
Probe whether PVD1/ESD1 frequency deadband is active on one dispatch.

This script creates a few temporary dynamic-case variants, replays the same
dispatch through TDS, and records both frequency and internal DG traces such as
deadband output and active-power command.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

import run_dispatch_tds as rdt


ROOT = rdt.ROOT
CASES = rdt.CASES
RESULTS = rdt.RESULTS
DEFAULT_RESULTS_DIR = RESULTS / "deadband_probe"


@dataclass(frozen=True)
class Variant:
    name: str
    ddn: float
    fdbd: float
    fdbdu: float
    note: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dispatch-json", type=Path, default=None,
                        help="Existing dispatch JSON to replay.")
    parser.add_argument("--hour", type=int, default=13)
    parser.add_argument("--dispatch", type=int, default=2)
    parser.add_argument("--opf-case", type=Path, default=rdt.DEFAULT_OPF_CASE)
    parser.add_argument("--dyn-case", type=Path, default=rdt.DEFAULT_DYN_CASE)
    parser.add_argument("--stable-dyn-case", type=Path, default=rdt.DEFAULT_STABLE_DYN_CASE)
    parser.add_argument("--curve-file", type=Path, default=rdt.DEFAULT_CURVE_FILE)
    parser.add_argument("--results-dir", type=Path, default=DEFAULT_RESULTS_DIR)
    parser.add_argument("--duration-seconds", type=int, default=900)
    parser.add_argument("--agc-interval", type=int, default=4)
    parser.add_argument("--kp", type=float, default=0.03)
    parser.add_argument("--ki", type=float, default=0.01)
    parser.add_argument("--init-mode", choices=("dispatch", "first"), default="first")
    parser.add_argument("--wind-prefix", action="append", default=None)
    parser.add_argument("--solar-prefix", action="append", default=None)
    parser.add_argument("--active-ddn", type=float, default=1.0,
                        help="Nonzero ddn used for the active deadband variants.")
    parser.add_argument("--deadband-hz", type=float, default=0.017,
                        help="Deadband magnitude in Hz used for probe variants.")
    return parser.parse_args()


def build_default_variants(active_ddn: float, deadband_hz: float) -> list[Variant]:
    return [
        Variant(
            name="current_off",
            ddn=0.0,
            fdbd=0.0,
            fdbdu=deadband_hz,
            note="Current migrated case: ddn=0, asymmetric threshold.",
        ),
        Variant(
            name="asym_on",
            ddn=active_ddn,
            fdbd=0.0,
            fdbdu=deadband_hz,
            note="Activate current asymmetric thresholding.",
        ),
        Variant(
            name="sym_on",
            ddn=active_ddn,
            fdbd=-deadband_hz,
            fdbdu=deadband_hz,
            note="Activate symmetric deadband for comparison.",
        ),
    ]


def write_variant_case(base_case: Path, out_dir: Path, variant: Variant) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    dst = out_dir / f"{base_case.stem}_{variant.name}.xlsx"
    wb = openpyxl.load_workbook(base_case)

    for sheet_name in ("PVD1", "ESD1"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        col_map = {str(name): i + 1 for i, name in enumerate(headers)}

        for field, value in (("fdbd", variant.fdbd), ("fdbdu", variant.fdbdu), ("ddn", variant.ddn)):
            if field not in col_map:
                continue
            col = col_map[field]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col, value=value)

    wb.save(dst)
    return dst


def _sum(values: np.ndarray | list[float]) -> float:
    arr = np.asarray(values, dtype=float).reshape(-1)
    return float(arr.sum()) if arr.size else 0.0


def _mean(values: np.ndarray | list[float]) -> float:
    arr = np.asarray(values, dtype=float).reshape(-1)
    return float(arr.mean()) if arr.size else 0.0


def snapshot(sa, t: float) -> dict[str, float]:
    row = {
        "time_s": float(t),
        "freq_dev_hz": float((sa.ACEc.f.v[0] - 1.0) * sa.config.freq),
    }

    for model_name in ("PVD1", "ESD1"):
        model = getattr(sa, model_name)
        key = model_name.lower()

        if model.n == 0:
            row[f"{key}_fdev_mean_hz"] = 0.0
            row[f"{key}_db_sum"] = 0.0
            row[f"{key}_pref_sum"] = 0.0
            row[f"{key}_pext_sum"] = 0.0
            row[f"{key}_psum_sum"] = 0.0
            row[f"{key}_pe_sum"] = 0.0
            continue

        row[f"{key}_fdev_mean_hz"] = _mean(model.Fdev.v)
        row[f"{key}_db_sum"] = _sum(model.DB_y.v)
        row[f"{key}_pref_sum"] = _sum(model.Pref.v)
        row[f"{key}_pext_sum"] = _sum(model.Pext.v)
        row[f"{key}_psum_sum"] = _sum(model.Psum.v)
        row[f"{key}_pe_sum"] = float(np.sum(np.asarray(model.Ipout_y.v, dtype=float) * np.asarray(model.v.v, dtype=float)))

    row["dg_db_sum"] = row["pvd1_db_sum"] + row["esd1_db_sum"]
    row["dg_pe_sum"] = row["pvd1_pe_sum"] + row["esd1_pe_sum"]
    row["dg_psum_sum"] = row["pvd1_psum_sum"] + row["esd1_psum_sum"]
    return row


def run_tds_trace(
    dispatch_record: rdt.DispatchRecord,
    curve: pd.DataFrame,
    dyn_case: Path,
    duration_seconds: int,
    agc_interval: int,
    kp: float,
    ki: float,
    wind_prefixes: tuple[str, ...],
    solar_prefixes: tuple[str, ...],
    init_mode: str,
) -> pd.DataFrame:
    rdt.validate_curve_window(curve, dispatch_record, duration_seconds)

    sa = rdt.andes.load(str(dyn_case), setup=False, no_output=True, default_config=True)
    sa.add("Output", dict(model="ACEc", varname="f"))
    sa.setup()

    link = rdt.build_andes_link(sa)

    pq_idx = sa.PQ.idx.v
    stg = sa.StaticGen.get_all_idxes()
    stg_w2t, stg_pv = rdt.pvd1_gen_subsets(sa, wind_prefixes, solar_prefixes)
    p0_w2t = sa.StaticGen.get(src="p0", attr="v", idx=stg_w2t)
    p0_pv = sa.StaticGen.get(src="p0", attr="v", idx=stg_pv)
    pvd1_w2t = sa.PVD1.find_idx(keys="gen", values=stg_w2t)
    pvd1_pv = sa.PVD1.find_idx(keys="gen", values=stg_pv)

    sap0 = sa.PQ.p0.v.copy()
    saq0 = sa.PQ.q0.v.copy()

    sa.StaticGen.set(src="p0", idx=dispatch_record.gen, attr="v", value=dispatch_record.pg)
    sa.Bus.set(src="v0", idx=dispatch_record.bus, attr="v", value=dispatch_record.vBus)
    sa.Bus.set(src="a0", idx=dispatch_record.bus, attr="v", value=dispatch_record.aBus)

    pv_bus = sa.PV.bus.v
    slack_bus = sa.Slack.bus.v
    sa.PV.set(src="v0", idx=sa.PV.idx.v, attr="v", value=sa.Bus.get(src="v0", attr="v", idx=pv_bus))
    sa.Slack.set(src="a0", idx=sa.Slack.idx.v, attr="v", value=sa.Bus.get(src="a0", attr="v", idx=slack_bus))

    stg_on = rdt.dispatch_online_mask(stg, dispatch_record)
    sn = sa.StaticGen.get(src="Sn", attr="v", idx=stg)
    bf = stg_on * sn / (stg_on * sn).sum()

    sa.PQ.config.p2p = 1
    sa.PQ.config.q2q = 1
    sa.PQ.config.p2z = 0
    sa.PQ.config.q2z = 0
    sa.PQ.pq2z = 0

    sa.TDS.config.criteria = 0
    sa.TDS.config.no_tqdm = True

    init_load, init_wind, init_solar = rdt.resolve_initial_profile(
        curve=curve,
        dispatch_record=dispatch_record,
        duration_seconds=duration_seconds,
        init_mode=init_mode,
    )

    sa.PQ.set(src="p0", idx=pq_idx, attr="v", value=init_load * sap0)
    sa.PQ.set(src="q0", idx=pq_idx, attr="v", value=init_load * saq0)
    sa.StaticGen.set(src="p0", idx=stg_w2t, attr="v", value=init_wind * p0_w2t)
    sa.StaticGen.set(src="p0", idx=stg_pv, attr="v", value=init_solar * p0_pv)

    sa.PFlow.run()
    if sa.exit_code != 0:
        raise RuntimeError(f"PFlow failed with exit_code={sa.exit_code}")

    _ = sa.TDS.init()
    if sa.exit_code != 0:
        raise RuntimeError(f"TDS init failed with exit_code={sa.exit_code}")

    pext_max = 999 * np.ones(sa.DG.n)
    if hasattr(sa, "ESD1") and sa.ESD1.n:
        ess_uid = sa.DG.idx2uid(sa.ESD1.idx.v)
        pext_max[ess_uid] = 999

    rows = [snapshot(sa, 0.0)]
    ace_integral = 0.0
    ace_raw = 0.0
    r0 = dispatch_record.hour * 3600 + dispatch_record.dispatch * duration_seconds

    for t in range(duration_seconds):
        for col, has_col in (("agov", "has_gov"), ("adg", "has_dg"), ("arg", "has_rg")):
            link[col] = ace_raw * bf * link[has_col] * link["gammap"]

        if t % agc_interval == 0 and t > 0:
            agov_to_set = {gov: agov for gov, agov in zip(link["gov_idx"], link["agov"]) if pd.notna(gov)}
            if agov_to_set:
                gov_idx = list(agov_to_set.keys())
                paux0_raw = np.array(list(agov_to_set.values()))
                gov_syn = sa.TurbineGov.get(src="syn", attr="v", idx=gov_idx)
                gov_gen = sa.SynGen.get(src="gen", attr="v", idx=gov_syn)
                gov_pmax = sa.StaticGen.get(src="pmax", attr="v", idx=gov_gen)
                gov_pmin = sa.StaticGen.get(src="pmin", attr="v", idx=gov_gen)
                gov_pref0 = sa.TurbineGov.get(src="pref0", attr="v", idx=gov_idx)
                gov_up = np.maximum(0.0, gov_pmax - gov_pref0)
                gov_dn = np.minimum(0.0, gov_pmin - gov_pref0)
                paux0 = np.where(
                    paux0_raw >= 0.0,
                    np.minimum(paux0_raw, gov_up),
                    np.maximum(paux0_raw, gov_dn),
                )
                sa.TurbineGov.set(src="paux0", idx=gov_idx, attr="v", value=paux0)

            adg_to_set = {dg: adg for dg, adg in zip(link["dg_idx"], link["adg"]) if pd.notna(dg)}
            if adg_to_set:
                dg_idx = list(adg_to_set.keys())
                pext0_raw = np.array(list(adg_to_set.values()))
                dg_uids = sa.DG.idx2uid(dg_idx)
                pext0 = np.minimum(pext0_raw, pext_max[dg_uids])
                sa.DG.set(src="Pext0", idx=dg_idx, attr="v", value=pext0)

        if t > 0:
            kload = curve["Load"].iloc[r0 + t]
            sa.PQ.set(src="Ppf", idx=sa.PQ.idx.v, attr="v", value=kload * sap0)
            sa.PQ.set(src="Qpf", idx=sa.PQ.idx.v, attr="v", value=kload * saq0)

            wind = curve["Wind"].iloc[r0 + t]
            sa.PVD1.set(src="pref0", idx=pvd1_w2t, attr="v", value=wind * p0_w2t)

            solar = curve["PV"].iloc[r0 + t]
            sa.PVD1.set(src="pref0", idx=pvd1_pv, attr="v", value=solar * p0_pv)

            sa.TDS.config.tf = t
            sa.TDS.run()

            ace_sum = sa.ACEc.ace.v.sum()
            ace_raw = -(kp * ace_sum + ki * ace_integral)
            ace_integral = ace_integral + ace_sum
            rows.append(snapshot(sa, float(sa.dae.t)))

        if sa.exit_code != 0:
            raise RuntimeError(f"TDS failed at t={t}s with exit_code={sa.exit_code}")

    return pd.DataFrame(rows)


def plot_variants(combined: pd.DataFrame, fig_path: Path, label: str) -> None:
    variants = list(combined["variant"].drop_duplicates())
    colors = {
        "current_off": "#0f5c78",
        "asym_on": "#b24c2a",
        "sym_on": "#2f7d32",
    }

    fig, axes = plt.subplots(4, 1, figsize=(13.5, 13.5), sharex=True)

    for variant in variants:
        df = combined[combined["variant"] == variant]
        color = colors.get(variant)
        axes[0].plot(df["time_s"], df["freq_dev_hz"], label=variant, linewidth=1.4, color=color)
        axes[1].plot(df["time_s"], df["dg_db_sum"], label=variant, linewidth=1.3, color=color)
        axes[2].plot(df["time_s"], df["dg_pe_sum"], label=variant, linewidth=1.3, color=color)
        axes[3].plot(df["time_s"], df["pvd1_db_sum"], label=f"{variant}:PVD1", linewidth=1.1, color=color)
        axes[3].plot(df["time_s"], df["esd1_db_sum"], label=f"{variant}:ESD1", linewidth=1.1, linestyle="--", color=color)

    axes[0].axhline(0.0, color="#777777", linestyle="--", linewidth=0.8)
    axes[1].axhline(0.0, color="#777777", linestyle="--", linewidth=0.8)
    axes[2].axhline(0.0, color="#777777", linestyle="--", linewidth=0.8)
    axes[3].axhline(0.0, color="#777777", linestyle="--", linewidth=0.8)

    axes[0].set_ylabel("Freq dev [Hz]")
    axes[0].set_title(f"Deadband probe for {label}")
    axes[1].set_ylabel("DG DB_y sum")
    axes[2].set_ylabel("DG Pe sum")
    axes[3].set_ylabel("PVD1 / ESD1 DB_y")
    axes[3].set_xlabel("Time [s]")

    for ax in axes:
        ax.grid(True, alpha=0.22)

    axes[0].legend(loc="upper right")
    axes[3].legend(loc="upper right", ncol=2, fontsize=9)
    fig.tight_layout()
    fig.savefig(fig_path, dpi=220)
    plt.close(fig)


def summarize_variant(df: pd.DataFrame, variant: Variant) -> dict[str, float | str]:
    freq = df["freq_dev_hz"]
    db = df["dg_db_sum"]
    pvd_db = df["pvd1_db_sum"]
    esd_db = df["esd1_db_sum"]

    return {
        "variant": variant.name,
        "note": variant.note,
        "ddn": variant.ddn,
        "fdbd_hz": variant.fdbd,
        "fdbdu_hz": variant.fdbdu,
        "samples": int(len(df)),
        "freq_min_hz": float(freq.min()),
        "freq_max_hz": float(freq.max()),
        "freq_abs_mean_hz": float(freq.abs().mean()),
        "db_abs_max": float(db.abs().max()),
        "db_nonzero_samples": int((db.abs() > 1e-9).sum()),
        "pvd1_db_abs_max": float(pvd_db.abs().max()),
        "esd1_db_abs_max": float(esd_db.abs().max()),
    }


def main() -> None:
    args = parse_args()
    rdt.andes.config_logger(stream_level=30)

    curve = rdt.load_curve(args.curve_file)
    base_stable = rdt.adapt_dyn_case(args.dyn_case, args.stable_dyn_case)
    variants = build_default_variants(args.active_ddn, args.deadband_hz)

    wind_prefixes = rdt.normalize_prefixes(args.wind_prefix, rdt.DEFAULT_WIND_PREFIXES)
    solar_prefixes = rdt.normalize_prefixes(args.solar_prefix, rdt.DEFAULT_SOLAR_PREFIXES)

    if args.dispatch_json is not None:
        dispatch_record = rdt.DispatchRecord.from_json(args.dispatch_json)
    else:
        import ams

        ams.config_logger(stream_level=50)
        dispatch_record = rdt.compute_dispatch(
            args.hour,
            args.dispatch,
            curve,
            args.opf_case,
            args.duration_seconds,
        )

    if not dispatch_record.converged:
        raise RuntimeError(f"Dispatch {dispatch_record.label} did not converge")

    out_dir = args.results_dir / dispatch_record.label
    variant_case_dir = out_dir / "cases"
    out_dir.mkdir(parents=True, exist_ok=True)

    summaries: list[dict[str, float | str]] = []
    all_frames: list[pd.DataFrame] = []

    for variant in variants:
        variant_case = write_variant_case(base_stable, variant_case_dir, variant)
        df = run_tds_trace(
            dispatch_record=dispatch_record,
            curve=curve,
            dyn_case=variant_case,
            duration_seconds=args.duration_seconds,
            agc_interval=args.agc_interval,
            kp=args.kp,
            ki=args.ki,
            wind_prefixes=wind_prefixes,
            solar_prefixes=solar_prefixes,
            init_mode=args.init_mode,
        )
        df.insert(0, "variant", variant.name)
        csv_path = out_dir / f"{dispatch_record.label}_{variant.name}_probe.csv"
        df.to_csv(csv_path, index=False)
        summaries.append(summarize_variant(df, variant))
        all_frames.append(df)

    combined = pd.concat(all_frames, ignore_index=True)
    combined_csv = out_dir / f"{dispatch_record.label}_deadband_probe_all.csv"
    summary_csv = out_dir / f"{dispatch_record.label}_deadband_probe_summary.csv"
    plot_png = out_dir / f"{dispatch_record.label}_deadband_probe.png"
    config_json = out_dir / f"{dispatch_record.label}_deadband_probe_config.json"

    combined.to_csv(combined_csv, index=False)
    pd.DataFrame(summaries).to_csv(summary_csv, index=False)
    plot_variants(combined, plot_png, dispatch_record.label)

    config = {
        "dispatch_label": dispatch_record.label,
        "agc_interval": args.agc_interval,
        "kp": args.kp,
        "ki": args.ki,
        "init_mode": args.init_mode,
        "variants": [variant.__dict__ for variant in variants],
        "base_stable_case": str(base_stable),
        "curve_file": str(args.curve_file),
    }
    config_json.write_text(json.dumps(config, indent=2))

    print(f"plot_png={plot_png}")
    print(f"summary_csv={summary_csv}")
    print(f"combined_csv={combined_csv}")
    print(pd.DataFrame(summaries).to_string(index=False))


if __name__ == "__main__":
    main()
